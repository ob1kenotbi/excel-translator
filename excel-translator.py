import os
import openpyxl
from deep_translator import GoogleTranslator

def translate_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    translator = GoogleTranslator(source="ja", target="en")

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    try:
                        cell.value = translator.translate(cell.value)
                    except Exception as e:
                        print(f"Error translating cell {cell.coordinate}: {e}")

    translated_filename = f"translated_{filename}"
    workbook.save(translated_filename)

    print(f"Translation completed and saved to {translated_filename}")

def read_excel():
    #start read excel
    for filename in os.listdir():
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            print(filename)
            #start translate
            translate_excel(filename)

def move_file():
    source = os.getcwd()
    destination_translated = os.path.join(source, "Translated")
    destination_archived = os.path.join(source, "Archive")

    for filename in os.listdir(source):
        if filename.startswith("translated"):
            src_path = os.path.join(source, filename)
            dst_path = os.path.join(destination_translated, filename)
            os.rename(src_path, dst_path)
            print(f"Finished moved {filename}")
        elif filename.endswith(".xlsx"):
            src_path = os.path.join(source, filename)
            dst_path = os.path.join(destination_archived, filename)
            os.rename(src_path, dst_path)
            print(f"Finished moved {filename}")

read_excel()
move_file()

